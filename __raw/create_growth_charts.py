import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Read the CSV
df = pd.read_csv('/Users/ricardopascua/Desktop/comparative_financial_ratios.csv')

# Create a new Excel workbook
writer = pd.ExcelWriter('/Users/ricardopascua/Desktop/growth_performance_analysis.xlsx', engine='openpyxl')

# Extract growth data
growth_rows = df[df['Metric'].str.contains('Sales|Operating income|Net income', na=False)]
growth_data = growth_rows.copy()

# Restructure data for better charting
airlines = [
    'Southwest Airlines', 'Continental Airlines', 'Lufthansa',
    'Singapore Airlines', 'Air Canada', 'WestJet Airlines'
]

# Create a summary sheet for growth metrics
summary_data = []
for _, row in growth_data.iterrows():
    metric = row['Metric']
    for airline in airlines:
        # Get most recent year data for each airline
        if airline == 'Southwest Airlines':
            value = row['Southwest Airlines 2008']
        elif airline == 'Continental Airlines':
            value = row['Continental Airlines 2008']
        elif airline == 'Lufthansa':
            value = row['Lufthansa 2008']
        elif airline == 'Singapore Airlines':
            value = row['Singapore Airlines 2009']
        elif airline == 'Air Canada':
            value = row['Air Canada 2008']
        elif airline == 'WestJet Airlines':
            value = row['WestJet Airlines 2008']

        # Convert percentage strings to float
        if isinstance(value, str) and '%' in value:
            value = float(value.replace('%', ''))

        summary_data.append({
            'Metric': metric,
            'Airline': airline,
            'Growth %': value
        })

summary_df = pd.DataFrame(summary_data)

# Create pivot table for charting
pivot_df = summary_df.pivot(index='Airline', columns='Metric', values='Growth %')

# Write to Excel
pivot_df.to_excel(writer, sheet_name='Growth Summary', startrow=1)
growth_data.to_excel(writer, sheet_name='Raw Data', index=False)

writer.close()

# Now load the workbook and add charts
wb = load_workbook('/Users/ricardopascua/Desktop/growth_performance_analysis.xlsx')
ws = wb['Growth Summary']

# Style the header
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)

for cell in ws[2]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Add title
ws['A1'] = 'Growth Performance - Competitive Analysis'
ws['A1'].font = Font(size=14, bold=True, color='366092')
ws.merge_cells('A1:D1')

# Create Bar Chart for Sales Growth
chart1 = BarChart()
chart1.type = 'col'
chart1.style = 10
chart1.title = 'Sales Growth % - Competitive Comparison'
chart1.y_axis.title = 'Growth %'
chart1.x_axis.title = 'Airlines'

data = Reference(ws, min_col=2, min_row=2, max_row=8, max_col=2)
cats = Reference(ws, min_col=1, min_row=3, max_row=8)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.height = 12
chart1.width = 20

ws.add_chart(chart1, 'F2')

# Create Bar Chart for Operating Income Growth
chart2 = BarChart()
chart2.type = 'col'
chart2.style = 11
chart2.title = 'Operating Income Growth % - Competitive Comparison'
chart2.y_axis.title = 'Growth %'
chart2.x_axis.title = 'Airlines'

data = Reference(ws, min_col=3, min_row=2, max_row=8, max_col=3)
cats = Reference(ws, min_col=1, min_row=3, max_row=8)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(cats)
chart2.height = 12
chart2.width = 20

ws.add_chart(chart2, 'F22')

# Create Bar Chart for Net Income Growth
chart3 = BarChart()
chart3.type = 'col'
chart3.style = 12
chart3.title = 'Net Income Growth % - Competitive Comparison'
chart3.y_axis.title = 'Growth %'
chart3.x_axis.title = 'Airlines'

data = Reference(ws, min_col=4, min_row=2, max_row=8, max_col=4)
cats = Reference(ws, min_col=1, min_row=3, max_row=8)
chart3.add_data(data, titles_from_data=True)
chart3.set_categories(cats)
chart3.height = 12
chart3.width = 20

ws.add_chart(chart3, 'F42')

# Create a combined comparison chart on new sheet
ws_combined = wb.create_sheet('Combined Growth Metrics')

# Copy the pivot data
for r_idx, row in enumerate(dataframe_to_rows(pivot_df, index=True, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        ws_combined.cell(row=r_idx+1, column=c_idx, value=value)

# Style header
for cell in ws_combined[2]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

ws_combined['A1'] = 'All Growth Metrics - Competitive Comparison'
ws_combined['A1'].font = Font(size=14, bold=True, color='366092')

# Create grouped bar chart with all metrics
chart4 = BarChart()
chart4.type = 'col'
chart4.style = 13
chart4.title = 'All Growth Metrics by Airline'
chart4.y_axis.title = 'Growth %'
chart4.x_axis.title = 'Airlines'
chart4.grouping = 'clustered'

data = Reference(ws_combined, min_col=2, min_row=2, max_row=8, max_col=4)
cats = Reference(ws_combined, min_col=1, min_row=3, max_row=8)
chart4.add_data(data, titles_from_data=True)
chart4.set_categories(cats)
chart4.height = 15
chart4.width = 25

ws_combined.add_chart(chart4, 'F2')

# Adjust column widths
for ws_temp in [ws, ws_combined]:
    ws_temp.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D']:
        ws_temp.column_dimensions[col].width = 15

wb.save('/Users/ricardopascua/Desktop/growth_performance_analysis.xlsx')

print("Excel file with growth performance charts created successfully!")
print("Location: /Users/ricardopascua/Desktop/growth_performance_analysis.xlsx")
